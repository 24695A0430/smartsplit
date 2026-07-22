import io
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from decimal import Decimal
from groups.models import Group, GroupMember
from expenses.models import Expense, ExpenseSplit
from payments.models import Payment
from django.contrib.auth.models import User
import datetime

# ReportLab imports for PDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# OpenPyXL imports for Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@login_required
def reports_index_view(request):
    group_id = request.GET.get('group')
    month = request.GET.get('month') # format YYYY-MM
    member_id = request.GET.get('member')

    # Get user's groups
    user_memberships = GroupMember.objects.filter(user=request.user).select_related('group')
    groups = [m.group for m in user_memberships]
    group_ids = [g.id for g in groups]

    expenses = Expense.objects.filter(group_id__in=group_ids).select_related('group', 'paid_by').order_by('-date')

    if group_id:
        expenses = expenses.filter(group_id=group_id)
    if month:
        try:
            year, m_val = map(int, month.split('-'))
            expenses = expenses.filter(date__year=year, date__month=m_val)
        except ValueError:
            pass
    if member_id:
        expenses = expenses.filter(Q(paid_by_id=member_id) | Q(splits__user_id=member_id)).distinct()

    # Get members of selected group for filtering
    members = []
    if group_id:
        sel_group = get_object_or_404(Group, id=group_id)
        members = [gm.user for gm in sel_group.members.all()]

    context = {
        'groups': groups,
        'members': members,
        'expenses': expenses[:30], # Limit preview on page
        'selected_group_id': int(group_id) if group_id else None,
        'selected_month': month,
        'selected_member_id': int(member_id) if member_id else None,
    }
    return render(request, 'reports/index.html', context)

@login_required
def export_pdf_view(request):
    group_id = request.GET.get('group')
    month = request.GET.get('month')
    member_id = request.GET.get('member')

    user_memberships = GroupMember.objects.filter(user=request.user).select_related('group')
    group_ids = [m.group_id for m in user_memberships]

    expenses = Expense.objects.filter(group_id__in=group_ids).select_related('group', 'paid_by').order_by('-date')

    title_suffix = "Global Expense Report"
    if group_id:
        g = get_object_or_404(Group, id=group_id)
        expenses = expenses.filter(group_id=group_id)
        title_suffix = f"Group Report - {g.name}"
    if month:
        try:
            year, m_val = map(int, month.split('-'))
            expenses = expenses.filter(date__year=year, date__month=m_val)
            title_suffix += f" ({month})"
        except ValueError:
            pass
    if member_id:
        m_user = get_object_or_404(User, id=member_id)
        expenses = expenses.filter(Q(paid_by_id=member_id) | Q(splits__user_id=member_id)).distinct()
        title_suffix += f" (Member: {m_user.username})"

    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=15
    )
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=20
    )
    th_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white
    )
    td_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )

    story.append(Paragraph(title_suffix, title_style))
    story.append(Paragraph(f"Generated on {datetime.date.today().strftime('%B %d, %Y')} | Requested by {request.user.username}", meta_style))
    story.append(Spacer(1, 10))

    # Build Table
    data = [[
        Paragraph('Date', th_style),
        Paragraph('Title', th_style),
        Paragraph('Category', th_style),
        Paragraph('Group', th_style),
        Paragraph('Paid By', th_style),
        Paragraph('Amount (₹)', th_style)
    ]]

    total_amount = Decimal('0.00')
    for exp in expenses:
        total_amount += exp.amount
        data.append([
            Paragraph(exp.date.strftime('%Y-%m-%d'), td_style),
            Paragraph(exp.title, td_style),
            Paragraph(exp.category, td_style),
            Paragraph(exp.group.name, td_style),
            Paragraph(exp.paid_by.username, td_style),
            Paragraph(f"{exp.amount}", td_style)
        ])

    data.append([
        Paragraph('<b>Total Summary</b>', td_style),
        Paragraph('', td_style),
        Paragraph('', td_style),
        Paragraph('', td_style),
        Paragraph('', td_style),
        Paragraph(f"<b>₹{total_amount}</b>", td_style)
    ])

    table = Table(data, colWidths=[65, 130, 80, 100, 85, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F1F5F9')),
        ('LINEABOVE', (0,-1), (-1,-1), 1, colors.HexColor('#94A3B8')),
    ]))

    story.append(table)
    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="smartsplit_report_{datetime.date.today()}.pdf"'
    return response

@login_required
def export_excel_view(request):
    group_id = request.GET.get('group')
    month = request.GET.get('month')
    member_id = request.GET.get('member')

    user_memberships = GroupMember.objects.filter(user=request.user).select_related('group')
    group_ids = [m.group_id for m in user_memberships]

    expenses = Expense.objects.filter(group_id__in=group_ids).select_related('group', 'paid_by').order_by('-date')

    title_suffix = "Global Expense Report"
    if group_id:
        g = get_object_or_404(Group, id=group_id)
        expenses = expenses.filter(group_id=group_id)
        title_suffix = f"Group Report - {g.name}"
    if month:
        try:
            year, m_val = map(int, month.split('-'))
            expenses = expenses.filter(date__year=year, date__month=m_val)
            title_suffix += f" ({month})"
        except ValueError:
            pass
    if member_id:
        expenses = expenses.filter(Q(paid_by_id=member_id) | Q(splits__user_id=member_id)).distinct()
        m_user = User.objects.get(id=member_id)
        title_suffix += f" (Member: {m_user.username})"

    # Generate Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses Summary"

    # Style definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    font_total = Font(name="Calibri", size=11, bold=True)
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_total = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Document Title Block
    ws.merge_cells("A1:F1")
    ws["A1"] = title_suffix
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated on {datetime.date.today().strftime('%Y-%m-%d')} | Requested by {request.user.username}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")
    ws["A2"].alignment = align_left
    ws.row_dimensions[2].height = 20

    ws.append([]) # Spacer row

    # Headers
    headers = ['Date', 'Title', 'Category', 'Group', 'Paid By', 'Amount (INR)']
    ws.append(headers)
    ws.row_dimensions[4].height = 24

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx == 1 else (align_right if col_idx == 6 else align_left)
        cell.border = thin_border

    # Data Rows
    start_row = 5
    total_amount = Decimal('0.00')
    for exp in expenses:
        total_amount += exp.amount
        row_data = [
            exp.date.strftime('%Y-%m-%d'),
            exp.title,
            exp.category,
            exp.group.name,
            exp.paid_by.username,
            float(exp.amount)
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 6:
                cell.alignment = align_right
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = align_left

    # Total Summary Row
    ws.append(['Total Summary', '', '', '', '', float(total_amount)])
    tot_row = ws.max_row
    ws.row_dimensions[tot_row].height = 22
    
    for col_idx in range(1, 7):
        cell = ws.cell(row=tot_row, column=col_idx)
        cell.font = font_total
        cell.fill = fill_total
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = align_left
        elif col_idx == 6:
            cell.alignment = align_right
            cell.number_format = '#,##0.00'

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Skip merged title cell for length check
        for cell in col[3:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Write response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="smartsplit_report_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response
